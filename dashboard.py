import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import base64
from pathlib import Path
from datetime import date, timedelta

from tools.sales import (
    get_sales_summary_by_seller, get_top_products, get_sales_by_month,
    get_pending_orders, get_top_products_by_customer, get_sellers,
)
from tools.customers import (
    get_top_customers, get_customers_by_country,
    get_customer_stats, get_sales_by_country, get_customers_from_country,
)
from tools.invoices import get_revenue_summary, get_overdue_summary

# ── Configuración ─────────────────────────────────────────────────────────────
_logo_path    = Path(__file__).parent / "logo.png"
_favicon_path = Path(__file__).parent / "favicon.ico"
_logo_b64  = base64.b64encode(_logo_path.read_bytes()).decode() if _logo_path.exists() else ""
_logo_src  = f"data:image/png;base64,{_logo_b64}"

# Streamlit acepta PIL Image como page_icon
from PIL import Image as _PIL
if _favicon_path.exists():
    _icon = _PIL.open(_favicon_path)
elif _logo_path.exists():
    _icon = _PIL.open(_logo_path)
else:
    _icon = "🍓"

st.set_page_config(
    page_title="Fruta de Andalucía — Dashboard",
    page_icon=_icon,  # type: ignore
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS profesional ───────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Ocultar elementos de Streamlit */
#MainMenu, footer, header { visibility: hidden; }

.block-container { padding-top: 1rem; padding-bottom: 1rem; }

/* Fuente global */
html, body, [class*="css"] { font-family: 'Inter', 'Segoe UI', sans-serif; }

/* Header corporativo */
.corp-header {
    background: linear-gradient(135deg, #1B5E20 0%, #2E7D32 60%, #388E3C 100%);
    border-radius: 12px;
    padding: 1.4rem 2rem;
    margin-bottom: 1.5rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    box-shadow: 0 4px 20px rgba(27,94,32,0.3);
}
.corp-header h1 {
    color: white; font-size: 1.7rem; font-weight: 700;
    margin: 0; letter-spacing: -0.3px;
}
.corp-header .subtitle {
    color: #A5D6A7; font-size: 0.85rem; margin-top: 4px;
}
.corp-header .badge {
    background: rgba(255,255,255,0.15); border-radius: 20px;
    padding: 6px 16px; color: white; font-size: 0.8rem;
}


/* Secciones */
.section-header {
    display: flex; align-items: center; gap: 10px;
    margin: 1.5rem 0 0.8rem 0;
    padding-bottom: 8px;
    border-bottom: 2px solid #E8F5E9;
}
.section-header span { font-size: 1.05rem; font-weight: 700; color: #1B5E20; }
.section-icon { font-size: 1.2rem; }

/* Sidebar */
[data-testid="stSidebar"] {
    background: #F1F8E9;
    border-right: 1px solid #C8E6C9;
}
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stTextInput label,
[data-testid="stSidebar"] .stDateInput label {
    color: #2E7D32; font-weight: 600; font-size: 0.82rem;
}

/* Botón descarga */
.stDownloadButton button {
    background: white !important;
    color: #2E7D32 !important;
    border: 1.5px solid #2E7D32 !important;
    border-radius: 8px !important;
    font-size: 0.8rem !important;
    padding: 4px 14px !important;
    transition: all 0.2s !important;
}
.stDownloadButton button:hover {
    background: #2E7D32 !important;
    color: white !important;
}

/* Tablas */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }

/* Separador */
hr { border-color: #E8F5E9 !important; margin: 1.2rem 0 !important; }
</style>
""", unsafe_allow_html=True)

# ── Helpers ───────────────────────────────────────────────────────────────────

def section(icon, title):
    st.markdown(f"""
    <div class="section-header">
        <span class="section-icon">{icon}</span>
        <span>{title}</span>
    </div>""", unsafe_allow_html=True)

def chart_layout(fig, height=380):
    fig.update_layout(
        height=height,
        margin=dict(l=10, r=130, t=10, b=10),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, Segoe UI, sans-serif", size=11, color="#424242"),
        legend=dict(bgcolor="rgba(0,0,0,0)"),
    )
    # cliponaxis solo aplica a trazas bar/scatter, no a choropleth
    fig.update_traces(cliponaxis=False, selector=dict(type="bar"))
    fig.update_traces(cliponaxis=False, selector=dict(type="scatter"))
    return fig

def _excel_bytes(df: pd.DataFrame, hoja: str = "Datos") -> bytes:
    VERDE = "1B5E20"; BLANCO = "FFFFFF"; GRIS = "F5F5F5"
    def _f(bold=False, size=10, color="212121"):
        return Font(bold=bold, size=size, color=color, name="Calibri")
    def _r(c): return PatternFill("solid", fgColor=c)
    def _b():
        s = Side(style="thin", color="BDBDBD")
        return Border(left=s, right=s, top=s, bottom=s)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = hoja
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A1:{chr(64+len(df.columns))}1")
    c = ws["A1"]
    c.value = f"Fruta de Andalucía S.C.A.  ·  {hoja}  ·  {date.today():%d/%m/%Y}"
    c.font = _f(bold=True, size=13, color=BLANCO); c.fill = _r(VERDE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=j, value=str(col))
        c.font = _f(bold=True, size=10, color=BLANCO); c.fill = _r("2E7D32")
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = _b()
        ws.column_dimensions[chr(64+j)].width = max(18, len(str(col))+4)
    for i, row in enumerate(df.itertuples(index=False), 3):
        ws.row_dimensions[i].height = 16
        fill = _r(BLANCO) if i % 2 == 0 else _r(GRIS)
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = _f(size=10); c.fill = fill; c.border = _b()
            c.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left", vertical="center")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ── Sidebar — Filtros ─────────────────────────────────────────────────────────
hoy = date.today()
RANGOS = {
    "Este año":       (date(hoy.year, 1, 1), hoy),
    "Este mes":       (date(hoy.year, hoy.month, 1), hoy),
    "Este trimestre": (date(hoy.year, ((hoy.month-1)//3)*3+1, 1), hoy),
    "Esta semana":    (hoy - timedelta(days=hoy.weekday()), hoy),
    "Hoy":            (hoy, hoy),
    "Personalizado":  (date(hoy.year, 1, 1), hoy),
}

@st.cache_data(ttl=300)
def _get_sellers(): return get_sellers()
lista_cp = _get_sellers()
opciones_cp = {"Todos": 0} | {v["nombre"]: v["id"] for v in lista_cp}

with st.sidebar:
    st.markdown("""
    <div style="padding:18px 4px 4px 4px;">
        <p style="color:#9E9E9E; font-size:0.7rem; font-weight:700;
                  text-transform:uppercase; letter-spacing:1.5px; margin:0;">
            Filtros
        </p>
    </div>
    """, unsafe_allow_html=True)

    cp_sel = st.selectbox("🏭 Centro de producción", list(opciones_cp.keys()))
    cp_id  = opciones_cp[cp_sel]
    st.markdown("---")
    rango_sel = st.selectbox("📅 Período", list(RANGOS.keys()))
    d_from, d_to = RANGOS[rango_sel]
    date_from = st.date_input("Desde", value=d_from)
    date_to   = st.date_input("Hasta", value=d_to)
    st.markdown("---")
    cli_filtro = st.text_input("👤 Filtrar por cliente", placeholder="Nombre del cliente...")
    st.markdown("---")
    st.caption(f"Actualizado {hoy:%d/%m/%Y}")
    if st.button("🔄 Limpiar caché"):
        st.cache_data.clear()
        st.rerun()

year_actual   = hoy.year
year_anterior = year_actual - 1

# ── Cache ─────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def _revenue(df, dt, sid):          return get_revenue_summary(date_from=df, date_to=dt, seller_id=sid)
@st.cache_data(ttl=300)
def _top_customers(df, dt, sid):    return get_top_customers(date_from=df, date_to=dt, limit=10, seller_id=sid)
@st.cache_data(ttl=300)
def _overdue():                     return get_overdue_summary()
@st.cache_data(ttl=300)
def _sellers(df, dt, cname):        return get_sales_summary_by_seller(date_from=df, date_to=dt, customer_name=cname)
@st.cache_data(ttl=300)
def _top_products(sid):             return get_top_products(limit=15, seller_id=sid)
@st.cache_data(ttl=300)
def _sales_month(y, sid):           return get_sales_by_month(year=y, seller_id=sid)
@st.cache_data(ttl=300)
def _pending(sid):                  return get_pending_orders(limit=100, seller_id=sid)
@st.cache_data(ttl=300)
def _countries():                   return get_customers_by_country()
@st.cache_data(ttl=300)
def _sales_by_country(df, dt, sid): return get_sales_by_country(date_from=df, date_to=dt, seller_id=sid)
@st.cache_data(ttl=300)
def _customers_from_country(c):     return get_customers_from_country(country_name=c)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="corp-header">
    <div style="display:flex; align-items:center; gap:20px;">
        <img src="{_logo_src}"
             style="height:60px; filter:brightness(0) invert(1); opacity:0.92;">
        <div>
            <div style="color:white; font-size:1.55rem; font-weight:700; letter-spacing:-0.3px;">
                Dashboard Comercial
            </div>
            <div class="subtitle">Fruta de Andalucía S.C.A.  ·  Datos en tiempo real de Odoo 18</div>
        </div>
    </div>
    <div class="badge">📅 {date.today():%d/%m/%Y}</div>
</div>
""", unsafe_allow_html=True)


# ── Carga de datos ────────────────────────────────────────────────────────────
df_str = str(date_from)
dt_str = str(date_to)

with st.spinner("Conectando con Odoo..."):
    revenue          = _revenue(df_str, dt_str, cp_id)
    top_customers    = _top_customers(df_str, dt_str, cp_id)
    overdue          = _overdue()
    sellers          = _sellers(df_str, dt_str, cli_filtro)
    top_products     = _top_products(cp_id)
    mes_actual       = _sales_month(year_actual, cp_id)
    mes_anterior     = _sales_month(year_anterior, cp_id)
    pending          = _pending(cp_id)
    countries        = _countries()
    sales_by_country = _sales_by_country(df_str, dt_str, cp_id)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 1 — KPIs
# ══════════════════════════════════════════════════════════════════════════════
pct_cobrado = revenue["total_cobrado"] / revenue["total_facturado"] * 100 if revenue["total_facturado"] else 0
pct_pdte    = 100 - pct_cobrado

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("🧾 Facturas emitidas",    f"{revenue['num_facturas']:,}",          f"período seleccionado")
k2.metric("💶 Total facturado",       f"{revenue['total_facturado']:,.0f} €")
k3.metric("✅ Cobrado",               f"{revenue['total_cobrado']:,.0f} €",    f"{pct_cobrado:.1f}%")
k4.metric("⏳ Pendiente de cobro",    f"{revenue['total_pendiente']:,.0f} €",  f"-{pct_pdte:.1f}%",  delta_color="inverse")
k5.metric("📋 Presupuestos abiertos", f"{len(pending)}",                       "sin confirmar")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 2 — Top clientes + Centros de producción
# ══════════════════════════════════════════════════════════════════════════════
col1, col2 = st.columns(2)

with col1:
    section("🏆", "Top 10 clientes por volumen")
    if top_customers:
        df_c = pd.DataFrame(top_customers)
        fig = px.bar(df_c, x="total_comprado", y="cliente", orientation="h",
                     text="total_comprado", color="total_comprado",
                     color_continuous_scale=["#C8E6C9","#1B5E20"],
                     labels={"total_comprado": "Importe (€)", "cliente": ""})
        fig.update_traces(texttemplate="%{text:,.0f} €", textposition="outside",
                          marker_line_width=0)
        fig = chart_layout(fig, 420)
        fig.update_layout(yaxis={"categoryorder":"total ascending"}, coloraxis_showscale=False)
        st.plotly_chart(fig, use_container_width=True)
        st.download_button("⬇️ Excel", _excel_bytes(df_c,"Top Clientes"),
            "top_clientes.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_c")

with col2:
    section("🏭", "Ventas por centro de producción")
    if sellers:
        df_s = pd.DataFrame(sellers)
        fig_s = px.bar(df_s, x="total_vendido", y="vendedor", orientation="h",
                       text="total_vendido", color="total_vendido",
                       color_continuous_scale=["#BBDEFB","#0D47A1"],
                       labels={"total_vendido": "Importe (€)", "vendedor": ""})
        fig_s.update_traces(texttemplate="%{text:,.0f} €", textposition="outside",
                            marker_line_width=0)
        fig_s = chart_layout(fig_s, 420)
        fig_s.update_layout(yaxis={"categoryorder":"total ascending"}, coloraxis_showscale=False)
        st.plotly_chart(fig_s, use_container_width=True)
        st.download_button("⬇️ Excel", _excel_bytes(df_s,"Centros de Producción"),
            "centros_produccion.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_s")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 3 — Comparativa mensual + Top productos
# ══════════════════════════════════════════════════════════════════════════════
col3, col4 = st.columns(2)

with col3:
    section("📅", f"Comparativa mensual {year_anterior} vs {year_actual}")
    meses_es = {"January":"Ene","February":"Feb","March":"Mar","April":"Abr",
                "May":"May","June":"Jun","July":"Jul","August":"Ago",
                "September":"Sep","October":"Oct","November":"Nov","December":"Dic"}

    def parse_months(data, year):
        rows = []
        for g in data:
            mes_raw = g["mes"]
            try:
                mes_num = int(mes_raw.split("/")[0]) if "/" in str(mes_raw) else pd.to_datetime(mes_raw).month
                mes_label = pd.Timestamp(year=year, month=mes_num, day=1).strftime("%B")
                rows.append({"mes": meses_es.get(mes_label, mes_label), "total": g["total"], "orden": mes_num})
            except Exception:
                pass
        return pd.DataFrame(rows)

    df_act = parse_months(mes_actual, year_actual)
    df_ant = parse_months(mes_anterior, year_anterior)

    if not df_act.empty or not df_ant.empty:
        all_m = pd.DataFrame({"orden": range(1,13),
                              "mes": ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]})
        da = all_m.merge(df_act.rename(columns={"total":str(year_actual)}), on=["orden","mes"], how="left").fillna(0)
        db = all_m.merge(df_ant.rename(columns={"total":str(year_anterior)}), on=["orden","mes"], how="left").fillna(0)
        df_comp = da.merge(db[["mes", str(year_anterior)]], on="mes")

        fig_c = go.Figure()
        fig_c.add_trace(go.Bar(name=str(year_anterior), x=df_comp["mes"],
                               y=df_comp[str(year_anterior)], marker_color="#90CAF9",
                               marker_line_width=0))
        fig_c.add_trace(go.Bar(name=str(year_actual), x=df_comp["mes"],
                               y=df_comp[str(year_actual)], marker_color="#2E7D32",
                               marker_line_width=0))
        fig_c.update_layout(barmode="group")
        fig_c = chart_layout(fig_c, 380)
        fig_c.update_layout(margin=dict(l=10,r=10,t=10,b=10),
                            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(fig_c, use_container_width=True)

with col4:
    section("📦", "Top 15 productos")
    if top_products:
        df_p = pd.DataFrame(top_products)
        fig_p = px.bar(df_p, x="importe_total", y="producto", orientation="h",
                       text="importe_total", color="importe_total",
                       color_continuous_scale=["#FFE0B2","#E65100"])
        fig_p.update_traces(texttemplate="%{text:,.0f} €", textposition="outside",
                            marker_line_width=0)
        fig_p = chart_layout(fig_p, 380)
        fig_p.update_layout(yaxis={"categoryorder":"total ascending"}, coloraxis_showscale=False)
        st.plotly_chart(fig_p, use_container_width=True)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 4 — Mapa + Estado de cobro
# ══════════════════════════════════════════════════════════════════════════════
col5, col6 = st.columns(2)

with col5:
    section("🗺️", "Ventas por país — clic para ver clientes")
    if sales_by_country:
        df_map = pd.DataFrame(sales_by_country)

        fig_map = go.Figure(go.Choropleth(
            locations=df_map["pais"],
            locationmode="country names",
            z=df_map["total_ventas"],
            colorscale=[
                [0.0, "#C8E6C9"],
                [0.2, "#66BB6A"],
                [0.5, "#2E7D32"],
                [1.0, "#1B5E20"],
            ],
            showscale=True,
            colorbar=dict(
                title=dict(text="Ventas (€)", font=dict(size=11)),
                thickness=12, len=0.6,
                tickformat=",.0f",
                bgcolor="rgba(255,255,255,0.7)",
                borderwidth=0,
            ),
            customdata=df_map[["num_pedidos", "total_ventas"]].values,
            hovertemplate=(
                "<b>%{location}</b><br>"
                "Ventas: %{customdata[1]:,.0f} €<br>"
                "Pedidos: %{customdata[0]}<extra></extra>"
            ),
            marker_line_color="white",
            marker_line_width=1.2,
        ))

        fig_map = chart_layout(fig_map, 400)
        fig_map.update_layout(
            margin=dict(l=0, r=0, t=0, b=0),
            geo=dict(
                showframe=False,
                showcoastlines=False,
                showland=True,      landcolor="#EEEEEE",
                showocean=True,     oceancolor="#EBF5FB",
                showlakes=False,
                showcountries=True, countrycolor="#BDBDBD",
                projection_type="natural earth",
                bgcolor="rgba(0,0,0,0)",
            ),
        )

        evento = st.plotly_chart(
            fig_map, use_container_width=True,
            on_select="rerun", selection_mode="points", key="mapa_paises"
        )

        # Clic en un país → tabla de sus clientes
        puntos = evento.selection.get("points", []) if evento and evento.selection else []
        if puntos:
            pais_clic = puntos[0].get("hovertext") or puntos[0].get("location", "")
            if pais_clic:
                with st.spinner(f"Cargando clientes de {pais_clic}..."):
                    clientes_pais = _customers_from_country(pais_clic)
                if clientes_pais:
                    st.markdown(f"**🌍 Clientes en {pais_clic}**")
                    df_cp = pd.DataFrame(clientes_pais)
                    df_cp.columns = ["Cliente", "Pedidos", "Total (€)"]
                    st.dataframe(df_cp, use_container_width=True, hide_index=True, height=200,
                                 column_config={"Total (€)": st.column_config.NumberColumn(format="%.2f €")})
        else:
            st.caption("💡 Haz clic en una burbuja para ver los clientes de ese país")

with col6:
    section("💰", "Estado de cobro")
    if revenue["por_estado_pago"]:
        labels_map = {"paid":"Cobrado","partial":"Cobro parcial",
                      "not_paid":"Pendiente","in_payment":"En proceso","reversed":"Revertido"}
        estados = {labels_map.get(k,k): v for k,v in revenue["por_estado_pago"].items()}
        COLORES = {"Cobrado":"#2E7D32","Cobro parcial":"#66BB6A","Pendiente":"#E53935",
                   "En proceso":"#1565C0","Revertido":"#757575"}
        fig_pie = go.Figure(go.Pie(
            labels=list(estados.keys()), values=list(estados.values()),
            hole=0.52, marker_colors=[COLORES.get(k,"#9E9E9E") for k in estados],
            textinfo="percent+label", hovertemplate="%{label}<br>%{value:,.0f} €<extra></extra>",
        ))
        fig_pie = chart_layout(fig_pie, 360)
        fig_pie.update_layout(margin=dict(l=10,r=10,t=10,b=10), showlegend=False)
        st.plotly_chart(fig_pie, use_container_width=True)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 5 — Tablas: Presupuestos + Facturas vencidas
# ══════════════════════════════════════════════════════════════════════════════
col7, col8 = st.columns(2)

with col7:
    section("📋", f"Presupuestos pendientes  ({len(pending)})")
    if pending:
        df_pend = pd.DataFrame(pending)
        df_pend.columns = ["Referencia","Cliente","Fecha","Centro producción","Estado","Total (€)"]
        busq = st.text_input("", placeholder="🔍  Buscar por cliente o referencia...", key="busq_pend",
                             label_visibility="collapsed")
        if busq:
            mask = df_pend.apply(lambda r: r.astype(str).str.contains(busq, case=False).any(), axis=1)
            df_pend = df_pend[mask]
        st.dataframe(df_pend, use_container_width=True, hide_index=True, height=280,
                     column_config={"Total (€)": st.column_config.NumberColumn(format="%.2f €")})
        st.download_button("⬇️ Excel", _excel_bytes(df_pend,"Presupuestos"),
            "presupuestos.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_pend")
    else:
        st.success("No hay presupuestos pendientes.")

with col8:
    section("⚠️", f"Clientes con facturas vencidas  ({len(overdue)})")
    if overdue:
        df_ov = pd.DataFrame(overdue)
        df_ov.columns = ["Cliente","Nº Facturas","Deuda (€)"]
        busq2 = st.text_input("", placeholder="🔍  Buscar cliente...", key="busq_ov",
                              label_visibility="collapsed")
        if busq2:
            df_ov = df_ov[df_ov["Cliente"].str.contains(busq2, case=False)]
        st.dataframe(df_ov, use_container_width=True, hide_index=True, height=280,
                     column_config={"Deuda (€)": st.column_config.NumberColumn(format="%.2f €")})
        st.download_button("⬇️ Excel", _excel_bytes(df_ov,"Facturas Vencidas"),
            "facturas_vencidas.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_ov")
    else:
        st.success("Sin facturas vencidas.")

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 6 — Buscador de cliente
# ══════════════════════════════════════════════════════════════════════════════
section("🔍", "Buscador de cliente")
query = st.text_input("", placeholder="Escribe el nombre del cliente (mín. 3 letras)...",
                      label_visibility="collapsed", key="buscador_cliente")

if query and len(query) >= 3:
    with st.spinner("Consultando Odoo..."):
        stats    = get_customer_stats(query)
        productos = get_top_products_by_customer(query, limit=8)

    if "error" in stats:
        st.warning(stats["error"])
    else:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("🛒 Pedidos confirmados", f"{stats['pedidos_confirmados']:,}")
        c2.metric("🧾 Facturas emitidas",   f"{stats['facturas_emitidas']:,}")
        c3.metric("💶 Total facturado",      f"{stats['total_facturado']:,.0f} €")
        c4.metric("⏳ Deuda pendiente",      f"{stats['deuda_pendiente']:,.0f} €")

        if productos:
            st.write("")
            section("📦", f"Top productos — {stats['cliente']}")
            df_pr = pd.DataFrame(productos)
            fig_pr = px.bar(df_pr, x="importe_total", y="producto", orientation="h",
                            text="importe_total", color="importe_total",
                            color_continuous_scale=["#E8EAF6","#283593"])
            fig_pr.update_traces(texttemplate="%{text:,.0f} €", textposition="outside",
                                 marker_line_width=0)
            fig_pr = chart_layout(fig_pr, 340)
            fig_pr.update_layout(yaxis={"categoryorder":"total ascending"},
                                 coloraxis_showscale=False)
            st.plotly_chart(fig_pr, use_container_width=True)
