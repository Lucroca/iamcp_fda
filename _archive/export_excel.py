"""
Exporta resumen de ventas por vendedor a Excel con formato profesional.
Uso: python export_excel.py [nombre_vendedor]
"""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from datetime import date
from odoo_client import odoo
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Parámetros ────────────────────────────────────────────────────────────────
vendedor = sys.argv[1] if len(sys.argv) > 1 else "miguelines"

# ── Colores corporativos ──────────────────────────────────────────────────────
VERDE_OSCURO  = "1B5E20"
VERDE_MEDIO   = "2E7D32"
VERDE_CLARO   = "C8E6C9"
VERDE_HEADER  = "388E3C"
GRIS_FILA_PAR = "F5F5F5"
BLANCO        = "FFFFFF"
NEGRO         = "212121"

# ── Datos de Odoo ─────────────────────────────────────────────────────────────
print(f"Consultando ventas de '{vendedor}'...")

groups = odoo.read_group(
    model="sale.order",
    domain=[
        ("user_id.name", "ilike", vendedor),
        ("state", "in", ["sale", "done"]),
    ],
    fields=["partner_id", "amount_total:sum", "id:count"],
    groupby=["partner_id"],
    orderby="amount_total desc",
)

if not groups:
    print(f"No se encontraron ventas para '{vendedor}'")
    sys.exit(1)

total_general = sum(g["amount_total"] for g in groups)
total_pedidos = sum(g["partner_id_count"] for g in groups)
nombre_vendedor = vendedor.capitalize()

# ── Helpers de estilo ─────────────────────────────────────────────────────────
def fuente(bold=False, size=11, color=NEGRO, name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)

def relleno(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def borde_fino():
    lado = Side(style="thin", color="BDBDBD")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def borde_medio():
    lado = Side(style="medium", color=VERDE_OSCURO)
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def centrado(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def derecha():
    return Alignment(horizontal="right", vertical="center")

def izquierda():
    return Alignment(horizontal="left", vertical="center")

FMT_EURO   = '#,##0.00 "€"'
FMT_NUMERO = '#,##0'
FMT_PCT    = '0.0"%"'

# ── Crear libro ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resumen de Ventas"
ws.sheet_view.showGridLines = False

# Ancho de columnas
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 10

# ── Fila 1-2: Cabecera empresa ─────────────────────────────────────────────────
ws.row_dimensions[1].height = 14
ws.row_dimensions[2].height = 36
ws.merge_cells("B2:E2")
ws["B2"].value        = "FRUTA DE ANDALUCÍA S.C.A."
ws["B2"].font         = fuente(bold=True, size=18, color=BLANCO)
ws["B2"].fill         = relleno(VERDE_OSCURO)
ws["B2"].alignment    = centrado()

ws.row_dimensions[3].height = 20
ws.merge_cells("B3:E3")
ws["B3"].value     = f"Resumen de Ventas por Vendedor — {nombre_vendedor}"
ws["B3"].font      = fuente(bold=True, size=13, color=BLANCO)
ws["B3"].fill      = relleno(VERDE_MEDIO)
ws["B3"].alignment = centrado()

ws.row_dimensions[4].height = 16
ws.merge_cells("B4:E4")
ws["B4"].value     = f"Generado el {date.today().strftime('%d/%m/%Y')}  ·  Datos en tiempo real de Odoo 18"
ws["B4"].font      = fuente(size=9, color=BLANCO)
ws["B4"].fill      = relleno(VERDE_HEADER)
ws["B4"].alignment = centrado()

# ── Fila 6-9: KPIs ────────────────────────────────────────────────────────────
ws.row_dimensions[6].height = 14
ws.row_dimensions[7].height = 30
ws.row_dimensions[8].height = 22
ws.row_dimensions[9].height = 14

kpis = [
    ("B", "Clientes",         len(groups),      FMT_NUMERO),
    ("C", "Total pedidos",    total_pedidos,     FMT_NUMERO),
    ("D", "Total vendido",    total_general,     FMT_EURO),
    ("E", "Ticket medio",     total_general / total_pedidos if total_pedidos else 0, FMT_EURO),
]

for col, titulo, valor, fmt in kpis:
    ws[f"{col}7"].value     = titulo
    ws[f"{col}7"].font      = fuente(bold=True, size=9, color=VERDE_OSCURO)
    ws[f"{col}7"].fill      = relleno(VERDE_CLARO)
    ws[f"{col}7"].alignment = centrado()
    ws[f"{col}7"].border    = borde_fino()

    ws[f"{col}8"].value        = valor
    ws[f"{col}8"].font         = fuente(bold=True, size=14, color=VERDE_OSCURO)
    ws[f"{col}8"].fill         = relleno(BLANCO)
    ws[f"{col}8"].alignment    = centrado()
    ws[f"{col}8"].number_format = fmt
    ws[f"{col}8"].border       = borde_fino()

# ── Fila 11: Encabezados tabla ────────────────────────────────────────────────
ws.row_dimensions[11].height = 24
headers = ["#", "Cliente", "Nº Pedidos", "Total vendido (€)", "% s/Total"]
cols    = ["A", "B", "C", "D", "E"]

for col, header in zip(cols, headers):
    cell            = ws[f"{col}11"]
    cell.value      = header
    cell.font       = fuente(bold=True, size=10, color=BLANCO)
    cell.fill       = relleno(VERDE_OSCURO)
    cell.alignment  = centrado()
    cell.border     = borde_medio()

# ── Filas de datos ────────────────────────────────────────────────────────────
for i, g in enumerate(groups, start=1):
    row    = 11 + i
    nombre = g["partner_id"][1] if g["partner_id"] else "Sin cliente"
    pedidos_cli = g["partner_id_count"]
    total_cli   = g["amount_total"]
    pct         = (total_cli / total_general * 100) if total_general else 0
    fill_color  = BLANCO if i % 2 == 0 else GRIS_FILA_PAR

    ws.row_dimensions[row].height = 18

    datos = [
        ("A", i,           FMT_NUMERO, centrado()),
        ("B", nombre,      "@",        izquierda()),
        ("C", pedidos_cli, FMT_NUMERO, centrado()),
        ("D", total_cli,   FMT_EURO,   derecha()),
        ("E", pct,         FMT_PCT,    centrado()),
    ]
    for col, valor, fmt, alin in datos:
        cell               = ws[f"{col}{row}"]
        cell.value         = valor
        cell.font          = fuente(size=10)
        cell.fill          = relleno(fill_color)
        cell.alignment     = alin
        cell.number_format = fmt
        cell.border        = borde_fino()

# ── Fila total ────────────────────────────────────────────────────────────────
fila_total = 11 + len(groups) + 1
ws.row_dimensions[fila_total].height = 22
ws.merge_cells(f"A{fila_total}:B{fila_total}")
ws[f"A{fila_total}"].value     = "TOTAL"
ws[f"A{fila_total}"].font      = fuente(bold=True, size=11, color=BLANCO)
ws[f"A{fila_total}"].fill      = relleno(VERDE_MEDIO)
ws[f"A{fila_total}"].alignment = centrado()
ws[f"A{fila_total}"].border    = borde_medio()

totales = [
    ("C", total_pedidos, FMT_NUMERO),
    ("D", total_general, FMT_EURO),
    ("E", 100.0,         FMT_PCT),
]
for col, valor, fmt in totales:
    cell               = ws[f"{col}{fila_total}"]
    cell.value         = valor
    cell.font          = fuente(bold=True, size=11, color=BLANCO)
    cell.fill          = relleno(VERDE_MEDIO)
    cell.alignment     = centrado()
    cell.number_format = fmt
    cell.border        = borde_medio()

# ── Inmovilizar paneles y zoom ─────────────────────────────────────────────────
ws.freeze_panes = "B12"
ws.sheet_view.zoomScale = 110

# ── Guardar ───────────────────────────────────────────────────────────────────
nombre_archivo = f"ventas_{vendedor.lower().replace(' ', '_')}_{date.today()}.xlsx"
ruta = os.path.join(os.path.dirname(__file__), nombre_archivo)
wb.save(ruta)
print(f"Excel guardado: {ruta}")
